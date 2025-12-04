import logging
from odoo import http, _
from odoo.http import request
from odoo.osv.expression import AND
from odoo.addons.point_of_sale.controllers.main import PosController
import io
import json

_logger = logging.getLogger(__name__)

try:
    import xlsxwriter
except Exception:
    xlsxwriter = None

try:
    import openpyxl
except Exception:
    openpyxl = None


class CustomPosController(PosController):

    @http.route(['/gm_pos/web', '/gm_pos/ui'], type='http', auth='user')
    def custom_pos_web(self, config_id=False, from_backend=False, **k):
        """Open a pos session for the given config - IDENTICAL to original POS
        
        This produces 100% identical UI as the default POS but with custom route
        
        :param debug: The debug mode to load the session in.
        :type debug: str.
        :param config_id: id of the config that has to be loaded.
        :type config_id: str.
        :returns: object -- The rendered pos session (identical to original).
        """
        is_internal_user = request.env.user._is_internal()
        pos_config = False
        if not is_internal_user:
            return request.not_found()
            
        domain = [
                ('state', 'in', ['opening_control', 'opened']),
                ('user_id', '=', request.session.uid),
                ('rescue', '=', False)
                ]
                
        if config_id and request.env['pos.config'].sudo().browse(int(config_id)).exists():
            domain = AND([domain,[('config_id', '=', int(config_id))]])
            pos_config = request.env['pos.config'].sudo().browse(int(config_id))
            
        pos_session = request.env['pos.session'].sudo().search(domain, limit=1)

        # The same POS session can be opened by a different user => search without restricting to
        # current user. Note: the config must be explicitly given to avoid fallbacking on a random
        # session.
        if not pos_session and config_id:
            domain = [
                ('state', 'in', ['opening_control', 'opened']),
                ('rescue', '=', False),
                ('config_id', '=', int(config_id)),
            ]
            pos_session = request.env['pos.session'].sudo().search(domain, limit=1)

        if not pos_config or not pos_config.active or pos_config.has_active_session and not pos_session:
            return request.redirect('/odoo/action-point_of_sale.action_client_pos_menu')

        if not pos_config.has_active_session:
            pos_config.open_ui()
            pos_session = request.env['pos.session'].sudo().search(domain, limit=1)

        # The POS only works in one company, so we enforce the one of the session in the context
        session_info = pos_session._update_session_info(request.env['ir.http'].session_info())
        context = {
            'from_backend': 1 if from_backend else 0,
            'use_pos_fake_tours': True if k.get('tours', False) else False,
            'session_info': session_info,
            'login_number': pos_session.with_company(pos_session.company_id).login(),
            'pos_session_id': pos_session.id,
            'pos_config_id': pos_session.config_id.id,
            'access_token': pos_session.config_id.access_token,
        }
        
        # Use custom template that's identical to original POS
        response = request.render('gm_pos_access_code.index_custom', context)
        response.headers['Cache-Control'] = 'no-store'
        return response

    @http.route('/gm_pos_access_code/export_xlsx', type='http', auth='user', methods=['POST'], csrf=False)
    def export_xlsx(self, **kwargs):
        if not xlsxwriter:
            return request.make_response('xlsxwriter not available', headers=[('Content-Type', 'text/plain')])
        try:
            payload = json.loads(request.httprequest.data.decode('utf-8') or '{}')
        except Exception:
            payload = {}
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {'in_memory': True})
        ws = wb.add_worksheet('Template')
        bold = wb.add_format({'bold': True})
        # Remove Access Code column - auto-generated
        headers = ['Booking Code', 'Booking From', 'Booking Price', 'Bottle Size', 'Validity Value', 'Validity Unit']
        for col, h in enumerate(headers):
            ws.write(0, col, h, bold)
        # current values row
        row = 1
        ws.write(row, 0, payload.get('booking_code') or '')
        ws.write(row, 1, payload.get('booking_from') or '')
        try:
            ws.write_number(row, 2, float(payload.get('booking_price') or 0))
        except Exception:
            ws.write(row, 2, payload.get('booking_price') or '')
        ws.write(row, 3, payload.get('bottle_size') or '100ml')
        try:
            ws.write_number(row, 4, int(payload.get('validity_value') or 7))
        except Exception:
            ws.write(row, 4, payload.get('validity_value') or 7)
        ws.write(row, 5, payload.get('validity_unit') or 'days')
        wb.close()
        output.seek(0)
        filename = 'booking_example.xlsx'
        headers = [
            ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', f'attachment; filename="{filename}"')
        ]
        return request.make_response(output.read(), headers=headers)

    @http.route('/gm_pos_access_code/import_xlsx', type='http', auth='user', methods=['POST'], csrf=False)
    def import_xlsx(self, **kwargs):
        if not openpyxl:
            return request.make_response(json.dumps({'error': 'openpyxl not available'}), headers=[('Content-Type', 'application/json')])
        file = request.httprequest.files.get('file')
        if not file:
            return request.make_response(json.dumps({'error': 'No file uploaded'}), headers=[('Content-Type', 'application/json')])
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            # Expect header row
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            headers_raw = [cell.value or '' for cell in header_cells]
            norm = lambda s: (str(s or '').strip().lower())
            headers = [norm(h) for h in headers_raw]
            header_index = {h: idx for idx, h in enumerate(headers)}
            def has(col):
                return norm(col) in header_index
            def idx(col):
                return header_index.get(norm(col))
            required = ['booking code', 'booking from', 'booking price', 'bottle size']
            missing = [c for c in required if not has(c)]
            if missing:
                return request.make_response(json.dumps({'error': f"Missing column(s): {', '.join(missing)}"}), headers=[('Content-Type', 'application/json')])
            codes = []
            created = 0
            for row in ws.iter_rows(min_row=2):
                # skip empty rows
                row_vals = [cell.value for cell in row]
                if not any(v not in (None, '') for v in row_vals):
                    continue
                def getv(col):
                    i = idx(col)
                    return row[i].value if i is not None else None
                # parse values safely
                booking_price = getv('booking price')
                try:
                    booking_price = float(booking_price or 0)
                except Exception:
                    booking_price = 0.0
                validity_value = getv('validity value')
                try:
                    validity_value = int(validity_value or 0)
                except Exception:
                    validity_value = 0
                validity_unit = str(getv('validity unit') or 'days')
                bottle_size = str(getv('bottle size') or '100ml')
                vals = {
                    'booking_code': getv('booking code') or '',
                    'booking_from': getv('booking from') or '',
                    'booking_price': booking_price,
                    'bottle_size': bottle_size,
                }
                if validity_value:
                    vals['validity_value'] = validity_value
                    vals['validity_unit'] = validity_unit
                # Access code auto-generated by model
                rec = request.env['pos.access.code'].sudo().create(vals)
                codes.append(rec.name)
                created += 1
            return request.make_response(json.dumps({'success': True, 'codes': codes, 'total': created}), headers=[('Content-Type', 'application/json')])
        except Exception as e:
            return request.make_response(json.dumps({'error': str(e)}), headers=[('Content-Type', 'application/json')])

    @http.route('/gm_pos_access_code/generate_pdf', type='http', auth='user', methods=['POST'], csrf=False)
    def generate_pdf(self, **kwargs):
        """Generate PDF with access codes using QWeb template"""
        try:
            import json
            data = json.loads(request.httprequest.data.decode('utf-8') or '{}')
            codes = data.get('codes', [])
            total_pages = data.get('totalPages', len(codes))
            company_logo_url = data.get('companyLogoUrl', '/gm_web/static/src/img/logo.png')
            
            if not codes:
                return request.make_response('No codes provided', headers=[('Content-Type', 'text/plain')])
            
            # Get company logo URL
            company = request.env.user.company_id
            company_logo_url = f'/web/image/res.company/{company.id}/logo' if company.logo else '/gm_web/static/src/img/logo.png'
            
            # Use QWeb template to generate PDF
            html_content = request.env['ir.ui.view']._render_template('gm_pos_access_code.access_code_template', {
                'codes': codes,
                'total_pages': total_pages,
                'company_logo_url': company_logo_url
            })
            
            headers = [
                ('Content-Type', 'text/html; charset=utf-8'),
                ('Content-Disposition', 'attachment; filename="access_codes.html"')
            ]
            return request.make_response(html_content, headers=headers)
            
        except Exception as e:
            return request.make_response(f'Error generating PDF: {str(e)}', headers=[('Content-Type', 'text/plain')])
